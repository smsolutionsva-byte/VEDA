"""Conservative tabular schedule detection + source-semantic MSPDI adaptation.

The original CSV/XLSX is the authoritative evidence source.  MSPDI is only a
transport format for Horizun: values synthesized for interoperability are never
allowed to become source facts in VEDA.
"""
from __future__ import annotations

import csv
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import Element, SubElement, ElementTree, register_namespace

SUPPORTED = {".csv", ".tsv", ".xlsx", ".xlsm"}

ALIASES = {
    "id": {"id", "uid", "unique id", "unique_id", "activity id", "activity_id", "activity code", "activity_code", "task id", "task_id"},
    "name": {"name", "task name", "task_name", "activity name", "activity_name", "description", "activity description", "activity_description"},
    "start": {"start", "start date", "start_date", "planned start", "planned_start", "target start", "target_start"},
    "finish": {"finish", "finish date", "finish_date", "end", "end date", "end_date", "planned finish", "planned_finish", "target finish", "target_finish"},
    "baseline_start": {"baseline start", "baseline_start", "baseline start date", "baseline_start_date"},
    "baseline_finish": {"baseline finish", "baseline_finish", "baseline finish date", "baseline_finish_date"},
    "actual_start": {"actual start", "actual_start", "actualstart"},
    "actual_finish": {"actual finish", "actual_finish", "actualfinish"},
    "pct": {"% complete", "percent complete", "percent_complete", "pct complete", "pct_complete", "progress", "physical %", "physical_percent"},
    "status": {"status", "activity status", "activity_status", "task status", "task_status", "state"},
    "wbs": {"wbs", "wbs path", "wbs_path", "wbs code", "wbs_code", "parent wbs", "parent_wbs"},
    "level": {"outline level", "outline_level", "level", "wbs level", "wbs_level"},
    "pred": {"predecessors", "predecessor", "pred", "preds", "predecessor ids", "predecessor_ids"},
    "summary": {"summary", "is summary", "is_summary"},
    "milestone": {"milestone", "is milestone", "is_milestone"},
    "activity_type": {"activity type", "activity_type", "task type", "task_type", "type"},
    "duration": {"duration", "duration days", "duration_days", "planned duration days", "planned_duration_days", "planned duration", "planned_duration"},
    "data_date": {"data date", "data_date", "status date", "status_date", "as of", "as_of"},
    "forecast_finish": {"forecast finish", "forecast_finish", "current finish", "current_finish", "remaining early finish", "remaining_early_finish"},
    "calendar": {"calendar", "calendar name", "calendar_name"},
    "resource": {"resource", "resource name", "resource_name", "resource assigned", "resource_assigned", "assigned resource", "assigned_resource"},
    "budgeted_units": {"budgeted units", "budgeted_units", "planned units", "planned_units", "work", "budgeted work", "budgeted_work"},
    "total_float": {"total float", "total_float", "total float days", "total_float_days"},
    "free_float": {"free float", "free_float", "free float days", "free_float_days"},
    "critical": {"critical", "is critical", "is_critical"},
    "constraint_type": {"constraint", "constraint type", "constraint_type", "primary constraint", "primary_constraint"},
    "constraint_date": {"constraint date", "constraint_date"},
}

NEGATIVE_NAME_HINTS = (
    "milestone_register", "milestone register", "resource_allocation", "resource allocation",
    "calendar_def", "calendar definition", "dictionary", "earned_value", "earned value",
    "delay_analysis", "delay analysis", "import_template", "import template", "update_template",
)
SCHEDULE_NAME_HINTS = ("schedule", "programme", "program", "baseline", "master plan", "lookahead", "look-ahead")
ALT_HINTS = ("extended", "extension", "recovery", "alternate", "alternative", "draft", "what-if", "whatif")


def _norm(v: Any) -> str:
    s = str(v or "").strip().casefold().replace("\n", " ")
    s = re.sub(r"[\-_]+", "_", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _read_tables(path: str, ext: str) -> list[tuple[str, list[str], list[list[Any]]]]:
    ext = ext.lower()
    if ext in (".csv", ".tsv"):
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            sample = f.read(8192); f.seek(0)
            delim = "\t" if ext == ".tsv" else ","
            if ext == ".csv":
                try: delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
                except Exception: pass
            rows = list(csv.reader(f, delimiter=delim))
        if not rows: return []
        return [("(csv)", [str(v or "") for v in rows[0]], rows[1:])]
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = []
        try:
            for ws in wb.worksheets:
                rows = []
                for r in ws.iter_rows(values_only=True):
                    vals = list(r)
                    if any(str(v or "").strip() for v in vals): rows.append(vals)
                    if len(rows) >= 100000: break
                if rows: out.append((ws.title, [str(v or "") for v in rows[0]], rows[1:]))
        finally:
            wb.close()
        return out
    return []


def _mapping(headers: list[Any]) -> dict[str, int]:
    hs = [_norm(h) for h in headers]
    out: dict[str, int] = {}
    for role, aliases in ALIASES.items():
        norms = {_norm(a) for a in aliases}
        for i, h in enumerate(hs):
            if h in norms:
                out[role] = i; break
    return out


def inspect(path: str, relative_path: str | None = None) -> dict:
    ext = os.path.splitext(path)[1].lower()
    if ext not in SUPPORTED:
        return {"candidate": False, "format": ext.lstrip(".")}
    display = (relative_path or os.path.basename(path)).replace("\\", "/")
    lowname = display.casefold()
    if any(x in lowname for x in NEGATIVE_NAME_HINTS):
        return {"candidate": False, "format": ext.lstrip("."), "reason": "known supporting/reference table"}
    best = None
    for sheet, headers, rows in _read_tables(path, ext):
        mp = _mapping(headers)
        core = sum(1 for k in ("id", "name", "start", "finish") if k in mp)
        dates = sum(1 for k in ("start", "finish") if k in mp)
        structure = sum(1 for k in ("wbs", "level", "pred") if k in mp)
        name_hint = any(x in lowname or x in sheet.casefold() for x in SCHEDULE_NAME_HINTS)
        candidate = core == 4 or (name_hint and "name" in mp and dates == 2 and structure >= 1)
        score = core * 20 + dates * 5 + structure * 4 + (8 if name_hint else 0) + min(len(rows), 200) / 200
        rec = {"candidate": candidate, "format": ext.lstrip("."), "sheet": sheet,
               "headers": headers, "mapping": mp, "row_count": len(rows), "score": score,
               "relative_path": display, "alternate_hint": any(x in lowname for x in ALT_HINTS)}
        if best is None or rec["score"] > best["score"]: best = rec
    return best or {"candidate": False, "format": ext.lstrip("."), "reason": "empty tabular file"}


def _parse_dt(v: Any) -> datetime | None:
    if isinstance(v, datetime): return v.replace(tzinfo=None)
    s = str(v or "").strip()
    if not s: return None
    s2 = s.replace("Z", "+00:00")
    for cand in (s2, s2.replace(" ", "T", 1)):
        try: return datetime.fromisoformat(cand).replace(tzinfo=None)
        except Exception: pass
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%b-%y"):
        try: return datetime.strptime(s, fmt)
        except Exception: pass
    return None


def _val(row: list[Any], mp: dict[str, int], role: str) -> Any:
    i = mp.get(role)
    return row[i] if i is not None and i < len(row) else None


def _bool(v: Any) -> bool | None:
    s = str(v or "").strip().casefold()
    if s in {"1", "true", "yes", "y", "x"}: return True
    if s in {"0", "false", "no", "n"}: return False
    return None


def _num(v: Any) -> float | None:
    try:
        s = str(v or "").strip().replace(",", "").replace("%", "")
        return float(s) if s else None
    except Exception:
        return None


def _pct_value(v: Any) -> float | None:
    x = _num(v)
    if x is None: return None
    if 0 <= x <= 1 and "%" not in str(v): x *= 100
    return max(0.0, min(100.0, x))


def _status(v: Any) -> str | None:
    s = str(v or "").strip().casefold().replace("_", " ")
    if not s: return None
    if s in {"not started", "notstart", "pending", "planned", "open"}: return "not_started"
    if s in {"in progress", "in_progress", "active", "started", "progress"}: return "in_progress"
    if s in {"complete", "completed", "finished", "done", "closed"}: return "complete"
    return None


def _activity_type(v: Any, explicit_summary: bool | None, explicit_milestone: bool | None) -> tuple[str, bool, bool]:
    s = str(v or "").strip()
    low = s.casefold().replace("_", " ")
    summary = bool(explicit_summary) or low in {"wbs summary", "summary", "summary task"}
    milestone = bool(explicit_milestone) or "milestone" in low
    if summary: canonical = "WBS Summary"
    elif "finish milestone" in low: canonical = "Finish Milestone"
    elif "start milestone" in low: canonical = "Start Milestone"
    elif "level of effort" in low or low == "loe": canonical = "Level of Effort"
    elif "resource dependent" in low: canonical = "Resource Dependent"
    elif "task dependent" in low: canonical = "Task Dependent"
    elif low in {"task", "activity", ""}: canonical = "Task"
    else: canonical = s
    return canonical, summary, milestone


def _split_resources(v: Any) -> list[str]:
    s = str(v or "").strip()
    if not s: return []
    # Commas often belong to resource names; semicolon/pipe are safer multi-value separators.
    vals = [x.strip() for x in re.split(r"[;|]+", s) if x.strip()]
    return vals or [s]


def _parse_pred_relations(v: Any, known_ids: set[str]) -> list[dict]:
    s = str(v or "").strip()
    if not s: return []
    out = []
    for raw in re.split(r"[;,|]+", s):
        tok = raw.strip()
        if not tok: continue
        if tok in known_ids:
            out.append({"pred_id": tok, "type": None, "lag_days": None,
                        "type_explicit": False, "lag_explicit": False})
            continue
        match = None
        for pid in sorted(known_ids, key=len, reverse=True):
            if not tok.startswith(pid): continue
            suffix = tok[len(pid):].strip()
            m = re.fullmatch(r"(?i)(FS|SS|FF|SF)(?:\s*([+-])\s*(\d+(?:\.\d+)?)\s*([dhw]))?", suffix)
            if m:
                lag = None
                lag_explicit = m.group(2) is not None
                if lag_explicit:
                    n = float(m.group(3)); unit = m.group(4).lower()
                    lag = n if unit == "d" else n / 8.0 if unit == "h" else n * 5.0
                    if m.group(2) == "-": lag = -lag
                else:
                    lag = 0.0
                match = {"pred_id": pid, "type": m.group(1).upper(), "lag_days": lag,
                         "type_explicit": True, "lag_explicit": lag_explicit}
                break
        if match: out.append(match)
        else:
            out.append({"pred_id": tok, "type": None, "lag_days": None,
                        "type_explicit": False, "lag_explicit": False})
    return out


def _wbs_nodes(paths: set[str]) -> list[dict]:
    all_wbs: set[str] = set()
    for path_value in paths:
        parts = [x.strip() for x in re.split(r"[./\\>]", path_value) if x.strip()]
        for i in range(1, len(parts) + 1): all_wbs.add(".".join(parts[:i]))
    out = []
    for code in sorted(all_wbs, key=lambda x: (x.count("."), x)):
        parent = code.rsplit(".", 1)[0] if "." in code else None
        out.append({"wbs_id": code, "parent_wbs_id": parent, "code": code,
                    "short_code": code.rsplit(".", 1)[-1], "name": code.rsplit(".", 1)[-1],
                    "parent_code": parent, "level": code.count(".") + 1, "project_node": False})
    return out


def inspect_semantics(path: str, relative_path: str | None = None) -> dict:
    """Inspect one tabular schedule without trusting any transport defaults."""
    meta = inspect(path, relative_path)
    if not meta.get("candidate"):
        return {**meta, "source_guarded": False}
    ext = os.path.splitext(path)[1].lower()
    tables = _read_tables(path, ext)
    table = next((x for x in tables if x[0] == meta.get("sheet")), tables[0])
    sheet, headers, rows = table
    mp = _mapping(headers)
    budget_header = (_norm(headers[mp["budgeted_units"]])
                     if "budgeted_units" in mp and mp["budgeted_units"] < len(headers) else "")
    # Microsoft Project's Work element is a duration quantity. Generic columns
    # named Budgeted_Units/Planned_Units are not assumed to be labor time.
    resource_work_values_compatible = budget_header in {
        _norm("work"), _norm("budgeted work"), _norm("budgeted_work")}
    clean_rows = [r for r in rows if str(_val(r, mp, "name") or "").strip()]
    source_ids = [str(_val(r, mp, "id") or i).strip() for i, r in enumerate(clean_rows, 1)]
    uid_by_source = {sid: i + 1 for i, sid in enumerate(source_ids)}
    known_ids = set(source_ids)

    task_by_id: dict[str, dict] = {}
    by_uid: dict[str, dict] = {}
    task_to_wbs: dict[str, str] = {}
    paths: set[str] = set()
    rels: list[dict] = []
    assignments: list[dict] = []
    planned_starts: list[datetime] = []
    planned_finishes: list[datetime] = []
    baseline_starts: list[datetime] = []
    baseline_finishes: list[datetime] = []
    forecast_finishes: list[datetime] = []
    data_dates: list[datetime] = []
    type_counts: Counter[str] = Counter()
    calendar_labels: Counter[str] = Counter()
    resource_labels: Counter[str] = Counter()
    statuses: Counter[str] = Counter()
    explicit_pct_count = 0
    baseline_coverage = 0
    baseline_equals_planned = 0
    duplicate_groups: dict[tuple[str, str], list[str]] = defaultdict(list)
    unnamed: list[str] = []
    milestone_duration_uids: list[int] = []
    milestone_resource_uids: list[int] = []
    duration_values: list[tuple[int, float]] = []
    explicit_critical: dict[int, bool] = {}
    float_values = 0
    constraints_present = 0

    for idx, r in enumerate(clean_rows, 1):
        sid = source_ids[idx - 1]; uid = uid_by_source[sid]
        name = str(_val(r, mp, "name") or sid).strip()
        if not name: unnamed.append(sid)
        start = _parse_dt(_val(r, mp, "start")); finish = _parse_dt(_val(r, mp, "finish"))
        bstart = _parse_dt(_val(r, mp, "baseline_start")); bfinish = _parse_dt(_val(r, mp, "baseline_finish"))
        astart = _parse_dt(_val(r, mp, "actual_start")); afinish = _parse_dt(_val(r, mp, "actual_finish"))
        ffinish = _parse_dt(_val(r, mp, "forecast_finish")); ddate = _parse_dt(_val(r, mp, "data_date"))
        duration = _num(_val(r, mp, "duration"))
        pct = _pct_value(_val(r, mp, "pct"))
        st = _status(_val(r, mp, "status"))
        explicit_summary = _bool(_val(r, mp, "summary"))
        explicit_milestone = _bool(_val(r, mp, "milestone"))
        atype, is_summary, is_milestone = _activity_type(_val(r, mp, "activity_type"), explicit_summary, explicit_milestone)
        wbs = str(_val(r, mp, "wbs") or "").strip()
        cal = str(_val(r, mp, "calendar") or "").strip()
        resources = _split_resources(_val(r, mp, "resource"))
        budgeted = _num(_val(r, mp, "budgeted_units"))
        tf = _num(_val(r, mp, "total_float")); ff = _num(_val(r, mp, "free_float"))
        crit = _bool(_val(r, mp, "critical"))
        ctype = str(_val(r, mp, "constraint_type") or "").strip()
        cdate = _parse_dt(_val(r, mp, "constraint_date"))

        if start: planned_starts.append(start)
        if finish: planned_finishes.append(finish)
        if bstart: baseline_starts.append(bstart)
        if bfinish: baseline_finishes.append(bfinish)
        if bstart and bfinish:
            baseline_coverage += 1
            if start and finish and start == bstart and finish == bfinish: baseline_equals_planned += 1
        if ffinish: forecast_finishes.append(ffinish)
        if ddate: data_dates.append(ddate)
        if pct is not None: explicit_pct_count += 1
        if st: statuses[st] += 1
        if cal: calendar_labels[cal] += 1
        if tf is not None or ff is not None: float_values += 1
        if ctype or cdate: constraints_present += 1
        if crit is not None: explicit_critical[uid] = crit
        if duration is not None: duration_values.append((uid, duration))
        type_counts[atype] += 1
        if wbs:
            paths.add(wbs); task_to_wbs[str(uid)] = wbs
        duplicate_groups[(wbs.casefold(), name.casefold())].append(sid)
        if is_milestone and duration not in (None, 0): milestone_duration_uids.append(uid)
        if is_milestone and resources: milestone_resource_uids.append(uid)
        for rn in resources:
            resource_labels[rn] += 1
            assignments.append({"task_uid": uid, "source_id": sid, "resource_name": rn,
                                "budgeted_units": budgeted})

        relations = _parse_pred_relations(_val(r, mp, "pred"), known_ids)
        for rel in relations:
            puid = uid_by_source.get(rel["pred_id"])
            if puid:
                rels.append({"pred_uid": puid, "succ_uid": uid,
                             "pred_id": rel["pred_id"], "succ_id": sid,
                             "type": rel["type"], "lag_days": rel["lag_days"],
                             "type_explicit": rel["type_explicit"], "lag_explicit": rel["lag_explicit"]})

        raw = {"task_id": str(uid), "source_id": sid, "activity_type": atype,
               "task_type": atype, "name": name, "wbs": wbs,
               "is_summary": is_summary, "is_milestone": is_milestone,
               "status": st, "percent_complete": pct,
               "target_start_date": start.isoformat(sep=" ") if start else "",
               "target_end_date": finish.isoformat(sep=" ") if finish else "",
               "baseline_start_date": bstart.isoformat(sep=" ") if bstart else "",
               "baseline_end_date": bfinish.isoformat(sep=" ") if bfinish else "",
               "actual_start_date": astart.isoformat(sep=" ") if astart else "",
               "actual_finish_date": afinish.isoformat(sep=" ") if afinish else "",
               "forecast_finish_date": ffinish.isoformat(sep=" ") if ffinish else "",
               "duration_days": duration, "calendar": cal,
               "resource_names": resources, "budgeted_units": budgeted,
               "total_float_days": tf, "free_float_days": ff,
               "critical": crit, "constraint_type": ctype or None,
               "constraint_date": cdate.isoformat(sep=" ") if cdate else None}
        task_by_id[str(uid)] = raw
        by_uid[str(uid)] = {"source_id": sid, "name": name, "wbs": wbs,
                            "outline_level": None, "summary": is_summary,
                            "milestone": is_milestone, "activity_type": atype}

    pred_uids = {r["succ_uid"] for r in rels}
    succ_uids = {r["pred_uid"] for r in rels}
    all_uids = set(range(1, len(clean_rows) + 1))
    open_logic = sorted((all_uids - pred_uids) | (all_uids - succ_uids)) if "pred" in mp else []
    duplicate_affected = sorted({sid for vals in duplicate_groups.values() if len(vals) > 1 for sid in vals})
    dup_groups = [{"wbs": k[0], "name": k[1], "activity_ids": vals}
                  for k, vals in duplicate_groups.items() if len(vals) > 1]

    status_exact = bool(clean_rows) and sum(statuses.values()) == len(clean_rows)
    if explicit_pct_count == len(clean_rows) and clean_rows:
        progress_available = True; progress_basis = "explicit percent-complete column"
    elif status_exact and not statuses.get("in_progress") and set(statuses).issubset({"not_started", "complete"}):
        progress_available = True; progress_basis = "activity status (exact 0/100 states)"
    else:
        progress_available = False; progress_basis = "not available from source"

    progress_pct = None
    if progress_available and clean_rows:
        vals = []
        for uid, raw in ((int(k), v) for k, v in task_by_id.items()):
            p = raw.get("percent_complete")
            if p is None:
                p = 100.0 if raw.get("status") == "complete" else 0.0 if raw.get("status") == "not_started" else None
            if p is not None: vals.append((p, raw.get("duration_days") or 1.0))
        if vals:
            den = sum(w for _, w in vals) or len(vals)
            progress_pct = round(sum(p * w for p, w in vals) / den, 1)

    rel_type_available = bool(rels) and all(r["type_explicit"] for r in rels)
    lag_available = rel_type_available and all(r["lag_days"] is not None for r in rels)
    baseline_values = baseline_coverage > 0
    critical_available = bool(clean_rows) and len(explicit_critical) == len(clean_rows)
    high_duration_uids = [uid for uid, d in duration_values if d > 44]
    duration_tasks = [int(k) for k, v in task_by_id.items() if not v.get("is_milestone") and not v.get("is_summary") and (v.get("duration_days") or 0) > 0]
    resourced = {a["task_uid"] for a in assignments}
    unresourced_uids = [u for u in duration_tasks if u not in resourced]

    meta.update({
        "source_guarded": True, "format": ext.lstrip("."), "sheet": sheet,
        "activity_count": len(clean_rows), "task_count": len(clean_rows),
        "task_ids": list(task_by_id), "task_by_id": task_by_id, "by_uid": by_uid,
        "task_to_wbs": task_to_wbs, "wbs_nodes": _wbs_nodes(paths),
        "wbs_count": len(_wbs_nodes(paths)), "task_type_counts": dict(type_counts),
        "planned_start": min(planned_starts).date().isoformat() if planned_starts else None,
        "planned_finish": max(planned_finishes).date().isoformat() if planned_finishes else None,
        "baseline_start": min(baseline_starts).date().isoformat() if baseline_starts else None,
        "baseline_finish": max(baseline_finishes).date().isoformat() if baseline_finishes else None,
        "baseline_values_available": baseline_values,
        "baseline_coverage_count": baseline_coverage, "baseline_coverage_total": len(clean_rows),
        "baseline_equals_planned_count": baseline_equals_planned,
        "baseline_assigned": False,
        "baseline_basis": "embedded tabular baseline columns" if baseline_values else None,
        "forecast_finish": max(forecast_finishes).date().isoformat() if forecast_finishes else None,
        "forecast_basis": "tabular forecast/current finish column" if forecast_finishes else "not available in source",
        "forecast_values_available": bool(forecast_finishes),
        "data_date": max(data_dates).date().isoformat() if data_dates else None,
        "data_date_available": bool(data_dates),
        "progress_available": progress_available, "progress_basis": progress_basis,
        "source_progress_pct": progress_pct, "status_counts": dict(statuses),
        "percent_values_available": explicit_pct_count > 0,
        "criticality_available": critical_available,
        "criticality_basis": "explicit source critical flag" if critical_available else "not available in source",
        "critical_count_source": sum(1 for v in explicit_critical.values() if v) if critical_available else None,
        "float_values_available": float_values > 0,
        "total_float_complete": ("total_float" in mp and all(v.get("total_float_days") is not None for v in task_by_id.values())),
        "constraint_values_available": constraints_present > 0,
        "constraint_columns_available": ("constraint_type" in mp or "constraint_date" in mp),
        "duration_values_complete": ("duration" in mp and all(v.get("duration_days") is not None for v in task_by_id.values())),
        "actual_finish_values_available": any(bool(v.get("actual_finish_date")) for v in task_by_id.values()),
        "predecessor_values_available": "pred" in mp,
        "relationship_type_values_available": rel_type_available,
        "lag_values_available": lag_available,
        "source_relationships": rels, "relationship_count_source": len(rels),
        "source_assignments": assignments, "resource_assignment_count": len(assignments),
        "resource_labels": sorted(resource_labels), "resource_label_count": len(resource_labels),
        "resource_work_values_compatible": resource_work_values_compatible,
        "duration_activity_count": len(duration_tasks),
        "milestone_uids": [int(k) for k, v in task_by_id.items() if v.get("is_milestone")],
        "resource_values_available": "resource" in mp,
        "calendar_labels": sorted(calendar_labels), "calendar_values_available": "calendar" in mp,
        "open_logic_uids": open_logic,
        "duplicate_sibling_groups": dup_groups, "duplicate_sibling_affected_ids": duplicate_affected,
        "milestone_duration_uids": milestone_duration_uids,
        "milestone_resource_uids": milestone_resource_uids,
        "high_duration_uids": high_duration_uids,
        "unresourced_uids": unresourced_uids,
        "unnamed_ids": unnamed,
        "project_resolution": "single_tabular_schedule",
        "ev_eligible": bool(baseline_values and data_dates and progress_available),
    })
    return meta


def _iso_duration_hours(hours: float | None) -> str | None:
    if hours is None: return None
    return f"PT{max(0.0, hours):.3f}H0M0S"


def prepare_mspdi(path: str, cache_dir: str | None = None, relative_path: str | None = None) -> tuple[str, dict]:
    meta = inspect_semantics(path, relative_path)
    if not meta.get("candidate"):
        raise ValueError("tabular file is not confidently schedule-shaped")
    task_by_id = meta["task_by_id"]
    by_uid = meta["by_uid"]

    register_namespace("", "http://schemas.microsoft.com/project")
    ns = "{http://schemas.microsoft.com/project}"
    root = Element(ns + "Project")
    SubElement(root, ns + "SaveVersion").text = "14"
    SubElement(root, ns + "Name").text = os.path.basename(path)
    if meta.get("planned_start"): SubElement(root, ns + "StartDate").text = meta["planned_start"] + "T08:00:00"
    if meta.get("planned_finish"): SubElement(root, ns + "FinishDate").text = meta["planned_finish"] + "T17:00:00"
    SubElement(root, ns + "ScheduleFromStart").text = "1"

    tasks_el = SubElement(root, ns + "Tasks")
    for uid_s in sorted(task_by_id, key=lambda x: int(x)):
        raw = task_by_id[uid_s]; uid = int(uid_s)
        t = SubElement(tasks_el, ns + "Task")
        # Deliberately keep every source activity at outline level 1. WBS context
        # stays in WBS/source metadata; the transport must not invent summary tasks.
        for tag, value in (("UID", uid), ("ID", uid), ("Name", raw.get("name") or raw.get("source_id") or uid),
                           ("Type", 1), ("IsNull", 0), ("WBS", raw.get("wbs") or raw.get("source_id") or uid),
                           ("OutlineNumber", uid), ("OutlineLevel", 1),
                           ("Milestone", 1 if raw.get("is_milestone") else 0),
                           ("Summary", 1 if raw.get("is_summary") else 0)):
            SubElement(t, ns + tag).text = str(value)
        if raw.get("percent_complete") is not None:
            SubElement(t, ns + "PercentComplete").text = str(round(float(raw["percent_complete"])))
        elif raw.get("status") in {"not_started", "complete"}:
            SubElement(t, ns + "PercentComplete").text = "100" if raw["status"] == "complete" else "0"
        for tag, key in (("Start", "target_start_date"), ("Finish", "target_end_date"),
                         ("ActualStart", "actual_start_date"), ("ActualFinish", "actual_finish_date")):
            dt = _parse_dt(raw.get(key))
            if dt: SubElement(t, ns + tag).text = dt.isoformat(timespec="seconds")
        if raw.get("duration_days") is not None:
            # MSPDI duration uses Project Time. The exact working-calendar meaning
            # is not used as source truth; VEDA retains source duration separately.
            SubElement(t, ns + "Duration").text = _iso_duration_hours(float(raw["duration_days"]) * 8.0)
            SubElement(t, ns + "DurationFormat").text = "7"
        if raw.get("baseline_start_date") or raw.get("baseline_end_date"):
            bl = SubElement(t, ns + "Baseline"); SubElement(bl, ns + "Number").text = "0"
            bs = _parse_dt(raw.get("baseline_start_date")); bf = _parse_dt(raw.get("baseline_end_date"))
            if bs: SubElement(bl, ns + "Start").text = bs.isoformat(timespec="seconds")
            if bf: SubElement(bl, ns + "Finish").text = bf.isoformat(timespec="seconds")
            if raw.get("duration_days") is not None:
                SubElement(bl, ns + "Duration").text = _iso_duration_hours(float(raw["duration_days"]) * 8.0)
                SubElement(bl, ns + "DurationFormat").text = "7"

    # Persist source relationships into transport. Unspecified predecessor tokens
    # require an interoperability default (FS/0), but metadata records that those
    # semantics were not explicit so VEDA QA cannot treat the defaults as evidence.
    for rel in meta.get("source_relationships") or []:
        succ = tasks_el[int(rel["succ_uid"]) - 1]
        pl = SubElement(succ, ns + "PredecessorLink")
        SubElement(pl, ns + "PredecessorUID").text = str(rel["pred_uid"])
        type_map = {"FF": "0", "FS": "1", "SF": "2", "SS": "3"}
        SubElement(pl, ns + "Type").text = type_map.get(rel.get("type") or "FS", "1")
        SubElement(pl, ns + "CrossProject").text = "0"
        lag_days = rel.get("lag_days") if rel.get("lag_days") is not None else 0.0
        # Project XML LinkLag is tenths of a minute. 8h working day = 480 min.
        SubElement(pl, ns + "LinkLag").text = str(int(round(float(lag_days) * 4800)))
        SubElement(pl, ns + "LagFormat").text = "7"

    # Separate Resources and Assignments are part of MSPDI. Use deterministic
    # transport UIDs; VEDA still stores their original SOURCE_FILE provenance.
    resources = sorted(meta.get("resource_labels") or [])
    resource_uid = {name: i + 1 for i, name in enumerate(resources)}
    if resources:
        res_el = SubElement(root, ns + "Resources")
        for name in resources:
            r = SubElement(res_el, ns + "Resource")
            SubElement(r, ns + "UID").text = str(resource_uid[name])
            SubElement(r, ns + "ID").text = str(resource_uid[name])
            SubElement(r, ns + "Name").text = name
            SubElement(r, ns + "Type").text = "1"
        ass_el = SubElement(root, ns + "Assignments")
        for i, a in enumerate(meta.get("source_assignments") or [], 1):
            ass = SubElement(ass_el, ns + "Assignment")
            SubElement(ass, ns + "UID").text = str(i)
            SubElement(ass, ns + "TaskUID").text = str(a["task_uid"])
            SubElement(ass, ns + "ResourceUID").text = str(resource_uid[a["resource_name"]])
            if meta.get("resource_work_values_compatible") and a.get("budgeted_units") is not None:
                SubElement(ass, ns + "Work").text = _iso_duration_hours(float(a["budgeted_units"]))

    outdir = Path(cache_dir or (str(Path(path).parent / ".veda-adapted")))
    outdir.mkdir(parents=True, exist_ok=True)
    out = outdir / (Path(path).stem + ".mspdi.xml")
    ElementTree(root).write(out, encoding="utf-8", xml_declaration=True)
    meta["canonical_path"] = str(out)
    return str(out), meta


__all__ = ["SUPPORTED", "inspect", "inspect_semantics", "prepare_mspdi"]
